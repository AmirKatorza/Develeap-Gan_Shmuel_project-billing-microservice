from flask import Flask, render_template, request, jsonify, abort, send_file ,redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy import create_engine
import yaml
from openpyxl import load_workbook, Workbook
import requests
from datetime import datetime, timedelta
from io import BytesIO
from CreateBill import *


app = Flask(__name__)

# Configure database
db_config = yaml.load(open('./billing-db.yaml'), Loader=yaml.FullLoader)
DATABASE_URL = f"mysql+mysqlconnector://{db_config['mysql_user']} \
                :{db_config['mysql_password']}@{db_config['mysql_host']} \
                /{db_config['mysql_db']}"
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
db = SQLAlchemy(app)

class Trucks(db.Model):
    __tablename__ = 'Trucks'
    id = db.Column(db.String(10), primary_key=True)
    provider_id = db.Column(db.Integer)

class Provider(db.Model):
    __tablename__ = 'Provider'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)

class Rates(db.Model):
    __tablename__ = 'Rates'
    product_id = db.Column(db.String(50), primary_key=True)
    rate = db.Column(db.Integer, primary_key=True)
    scope = db.Column(db.String(50), primary_key=True)


# POST /provider
# creates a new provider record:
# - name - provider name. must be unique.
# Returns a unique provider id as json: { "id":<str>}
@app.route("/provider", methods=["POST"])
def add_provider():
    try:
        name = request.json.get("name")
        new_provider = Provider(name=name)
        db.session.add(new_provider)
        db.session.commit()
        return jsonify({"id": new_provider.id}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Provider name must be unique."}), 400


# PUT /provider/{id} can be used to update provider name
@app.route('/provider/<int:provider_id>', methods=['PUT'])
def update_provider(provider_id):
    new_name = request.json.get('name')
    if not new_name:
        return jsonify({"error": "New name not provided"}), 400
    provider = db.session.query(Provider).filter_by(id=provider_id).first()
    if not provider:
        return jsonify({"error": "Provider not found"}), 404
    provider.name = new_name
    db.session.commit()
    return jsonify({"message": "Provider updated successfully"})


# POST /rates
# - file=<filename>
# Will upload new rates from an excel file in "/in" folder. Rate excel has the following columns:
# - Product - a product id
# - Rate - integer (in agorot)
# - Scope - ALL or A provider id.
# The new rates over-write the old ones
# A scoped rate has higher precedence than an "ALL" rate
@app.route('/rates', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']    
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    if file:
        try:
            db.session.query(Rates).delete()
            wb = load_workbook(file)
            sheet = wb.active 
            print("workbook loaded")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_name, rate, scope = row
                new_product = Rates(product_id=product_name, rate=rate, scope=scope)
                db.session.add(new_product)
            db.session.commit()
            return jsonify({"message": "Data uploaded and updated successfully"}), 200
        except Exception as e:
            print("Error updating rates:", e)
            return jsonify({"error": "An error occurred while updating rates"}), 500


# GET /rates
# Will download a copy of the same excel that was uploaded using POST /rates
@app.route('/rates', methods=['GET'])
def download_rates_excel():
    try:
        rates = db.session.query(Rates).all()
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Rates'       
        sheet.append(['product_id', 'rate', 'scope'])

        # Add data to the sheet
        for rate in rates:
            sheet.append([rate.product_id, rate.rate, rate.scope])

        excel_buffer = BytesIO()
        wb.save(excel_buffer)

        # Prepare the response with Excel file data
        excel_buffer.seek(0)
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,  #
            download_name='rates.xlsx'  
        )
    except Exception as e:
        print("Error generating Excel:", e)
        return jsonify({"error": "An error occurred while generating Excel"}), 500


# POST /truck
# registers a truck in the system
# - provider - known provider id
# - id - the truck license plate
@app.route('/truck', methods=['POST'])
def register_truck():
    try:
        provider_id = request.json.get('provider')
        truck_id = request.json.get('id')

        provider = Provider.query.get(provider_id)
        if not provider:
            return jsonify({"error": "Provider not found"}), 404

        new_truck = Trucks(id=truck_id, provider_id=provider_id)
        db.session.add(new_truck)
        db.session.commit()

        return jsonify({"message": "Truck registered successfully"}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Truck with the same ID already exists"}), 400


# PUT /truck/{id} can be used to update provider id
@app.route('/truck/<string:truck_id>', methods=['PUT'])
def update_truck(truck_id):
    new_provider = request.json.get('provider_id')
    if not new_provider:
        return jsonify({"error": "new provider not provided"}), 400
    truck = db.session.query(Trucks).filter_by(id=truck_id).first()
    if not db.session.query(Trucks).filter_by(id=truck_id).first():
        return jsonify({"error": "Truck not found"}), 404
    if not db.session.query(Provider).filter_by(id=new_provider).first():
        return jsonify({"error": "Provider not found"}), 404
    truck.provider_id = new_provider
    db.session.commit()
    return jsonify({"message": "Truck updated successfully"})


# GET /truck/<id>?from=t1&to=t2
# - id is the truck license. 404 will be returned if non-existent
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now".
# Returns a json:
# { "id": <str>,
#   "tara": <int>, // last known tara in kg
#   "sessions": [ <id1>,...]
# }
@app.route('/truck/<id>', methods=['GET'])
def get_truck_info(id):
    truck = Trucks.query.get(id)
    if not truck:
        return jsonify({"error": "Truck not found"}), 404
    t1 = request.args.get('from', (datetime.today().replace(day=1, hour=0, minute=0, second=0) - timedelta(days=1)).strftime('%Y%m%d%H%M%S'))
    t2 = request.args.get('to', datetime.now().strftime('%Y%m%d%H%M%S'))
    weight_api_url = f'http://weight-team-container-name:8080/item/{id}' # TODO - needs to changesd according to weight team container name
    params = {
        'from': t1,
        'to': t2,
    }
    try:
        response = requests.get(weight_api_url, params=params)
    except:
        return jsonify({"error": "Truck not found"}), 404    
    return jsonify(response), 200   


# GET /bill/<id>?from=t1&to=t2
# - id is provider id
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now". 
# Returns a json:
# {
#   "id": <str>,
#   "name": <str>,
#   "from": <str>,
#   "to": <str>,
#   "truckCount": <int>,
#   "sessionCount": <int>,
#   "products": [
#     { "product":<str>,
#       "count": <str>, // number of sessions
#       "amount": <int>, // total kg
#       "rate": <int>, // agorot
#       "pay": <int> // agorot
#     },
#     ...
#   ],
#   "total": <int> // agorot
# }
@app.route('/bill/<id>', methods=['GET'])
def get_bill(id):
    
    t1 = request.args.get('from', (datetime.today().replace(day=1, hour=0, minute=0, second=0) - timedelta(days=1)).strftime('%Y%m%d%H%M%S'))
    t2 = request.args.get('to', datetime.now().strftime('%Y%m%d%H%M%S'))
    f = 'in'   
    
    params = {
        'from': t1,
        'to': t2,
        'filter': f
    }
    create_bill(id, params)

# GET /health
#  - By default returns "OK" and status 200 OK
#  - If system depends on external resources (e.g. db), and they are not available (e.g. "select 1;" fails ) then it should return "Failure" and 500 Internal Server Error
#  - Failure of "weight" system is not relevant, i.e. payment system is OK even if weight system is not
@app.route('/health')
def health_check():
    try:
        test_engine = create_engine(DATABASE_URL)
        connection = test_engine.connect()
        connection.close()
        print("Database connection is successful!")
        return jsonify({"message": "Connected!"}), 200  # HTTP 200 OK
    except Exception as e:
        print("Database connection error:", e)
        return jsonify({"error": "Failed to connect!"}), 500
        abort(500)  # HTTP 500 Internal Server Error


@app.route('/main/<string:address>', methods=['GET','POST','PUT'])
def html_index(address):
    if address == "index": return render_template('main.html')
    elif request.method == "GET":
        return render_template(f'{address}.html')
    else:
        if address == "post_provider": #Add http request
            return "<h1>yo</h1>"     
    return "<h1>hi</h1>"
        
    
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)